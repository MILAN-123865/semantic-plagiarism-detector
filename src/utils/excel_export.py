"""
src/utils/excel_export.py
-------------------------
Utility for exporting similarity matrices into styled Excel (.xlsx) workbooks
with conditional formatting matching the application's heatmap logic.
Supports both in-memory generation and managed temporary disk-file creation with automatic exit cleanup.
Also provides streaming CSV generation for memory-efficient exports of large datasets.
"""

import atexit
import csv
import io
import os
import re
import tempfile
from datetime import datetime, timezone
from typing import Generator

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

from openpyxl.utils import get_column_letter

from src.utils.export_sanitizer import (
    FORMULA_TRIGGER_PREFIXES,
    sanitize_spreadsheet_value,
)


def sanitize_sheet_title(title: str) -> str:
    """
    Sanitize a worksheet title to comply with Excel's naming rules.

    Excel worksheet titles:
    - Cannot exceed 31 characters
    - Cannot contain [, ], *, ?, :, /, or .
    """
    sanitized_title = re.sub(r"[\[\]\*\?:/\.]", "", str(title))
    sanitized_title = sanitized_title[:31]

    return sanitized_title or "Sheet"


def _create_managed_temp_file(suffix: str = ".xlsx", prefix: str = "temp_") -> str:
    """Helper to create a temporary file that is automatically deleted on exit."""
    fd, temp_path = tempfile.mkstemp(suffix=suffix, prefix=prefix)
    os.close(fd)

    def _cleanup():
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass

    atexit.register(_cleanup)
    return temp_path


def _truncate_title(title, max_length: int = 60) -> str:
    """
    Truncate a title to max_length characters, appending '...' if truncated.

    Args:
        title: The title to truncate. Coerced to str, so a non-string
            DataFrame index (e.g. an integer document ID) does not raise.
        max_length: Maximum length before truncation (default: 60)

    Returns:
        Truncated title with '...' suffix if original was longer
    """
    title = str(title)
    if len(title) <= max_length:
        return title
    return title[: max_length - 3] + "..."


def build_similarity_workbook(
    df: pd.DataFrame,
    threshold: float = 0.59,
    write_only: bool = False,
    low_threshold: float = 0.0,
    mid_threshold: float = 0.59,
    high_threshold: float = 1.0,
) -> Workbook:
    """Helper function that builds and styles the openpyxl Workbook.

    Args:
        df: Similarity matrix DataFrame with document labels as index and columns.
        threshold: Score threshold for conditional formatting color scale.
        write_only: If True, uses openpyxl write_only mode with ws.append() for
            memory-efficient streaming of large matrices. Defaults to False.
        low_threshold: Low breakpoint for the 3-color scale.
        mid_threshold: Mid breakpoint for the 3-color scale.
        high_threshold: High breakpoint for the 3-color scale.

    Returns:
        Workbook: Configured openpyxl Workbook instance.
    """
    # Older callers pass ``threshold`` as the yellow midpoint.
    if threshold != 0.59 and mid_threshold == 0.59:
        mid_threshold = threshold

    if write_only:
        wb = Workbook(write_only=True)
        wb.properties.title = "Semantic Plagiarism Similarity Report"
        wb.properties.creator = "Semantic Plagiarism Detector"
        wb.properties.created = datetime.now(timezone.utc)

        ws = wb.create_sheet(title="Similarity Matrix")

        header_fill = PatternFill(
            start_color="1F2937", end_color="1F2937", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        # Write header row
        header_row = []
        c0 = WriteOnlyCell(ws, value="Document")
        c0.fill = header_fill
        c0.font = header_font
        c0.alignment = header_align
        header_row.append(c0)

        for col_name in df.columns:
            truncated_name = _truncate_title(col_name)
            cell = WriteOnlyCell(
                ws, value=sanitize_spreadsheet_value(truncated_name)
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            if len(str(col_name)) > 60:
                cell.comment = Comment(
                    sanitize_spreadsheet_value(str(col_name)), "Excel Export"
                )
            header_row.append(cell)
        ws.append(header_row)

        # Write data rows
        for index_label, row in df.iterrows():
            row_cells = []
            truncated_label = _truncate_title(index_label)
            label_cell = WriteOnlyCell(
                ws, value=sanitize_spreadsheet_value(truncated_label)
            )
            label_cell.fill = header_fill
            label_cell.font = header_font
            if len(str(index_label)) > 60:
                label_cell.comment = Comment(
                    sanitize_spreadsheet_value(str(index_label)), "Excel Export"
                )
            row_cells.append(label_cell)

            for val in row:
                val_cell = WriteOnlyCell(ws, value=float(val))
                val_cell.number_format = "0.0%"
                val_cell.alignment = Alignment(horizontal="right")
                row_cells.append(val_cell)

            ws.append(row_cells)

        # Apply Conditional Formatting (3-Color Scale)
        max_row = len(df) + 1
        max_col = len(df.columns) + 1

        if max_row > 1 and max_col > 1:
            start_cell = "B2"
            end_col_letter = get_column_letter(max_col)
            end_cell = f"{end_col_letter}{max_row}"
            matrix_range = f"{start_cell}:{end_cell}"

            color_scale = ColorScaleRule(
                start_type="num",
                start_value=low_threshold,
                start_color="FFFFFF",  # White (low)
                mid_type="num",
                mid_value=mid_threshold,
                mid_color="FEF08A",  # Yellow (mid)
                end_type="num",
                end_value=high_threshold,
                end_color="EF4444",  # Red (high)
            )
            ws.conditional_formatting.add(matrix_range, color_scale)

        # Auto-adjust column widths
        max_index_len = max([len(str(idx)) for idx in df.index] + [len("Document")])
        ws.column_dimensions[get_column_letter(1)].width = max(
            min(max_index_len, 60) + 3, 12
        )

        for col_idx, col_name in enumerate(df.columns, start=2):
            col_len = min(len(str(col_name)), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                col_len + 3, 12
            )

        return wb

    # Default write_only=False (in-memory DOM)
    wb = Workbook()
    wb.properties.title = "Semantic Plagiarism Similarity Report"
    wb.properties.creator = "Semantic Plagiarism Detector"
    wb.properties.created = datetime.now(timezone.utc)

    ws = wb.active
    ws.title = sanitize_sheet_title("Similarity Matrix")

    # Write headers and index labels with truncated titles, preserving full names in comments.
    # Labels originate from uploaded filenames, so they are sanitized before
    # being written to prevent formula injection in the exported workbook.
    ws.cell(row=1, column=1, value="Document")
    for col_idx, col_name in enumerate(df.columns, start=2):
        truncated_name = _truncate_title(col_name)
        cell = ws.cell(
            row=1, column=col_idx, value=sanitize_spreadsheet_value(truncated_name)
        )
        # Add full title as comment if truncated
        if len(str(col_name)) > 60:
            cell.comment = Comment(
                sanitize_spreadsheet_value(str(col_name)), "Excel Export"
            )

    for row_idx, (index_label, row) in enumerate(df.iterrows(), start=2):
        truncated_label = _truncate_title(index_label)
        cell = ws.cell(
            row=row_idx, column=1, value=sanitize_spreadsheet_value(truncated_label)
        )
        # Add full title as comment if truncated
        if len(str(index_label)) > 60:
            cell.comment = Comment(
                sanitize_spreadsheet_value(str(index_label)), "Excel Export"
            )

        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")

    # Header styling
    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font

    # Apply Conditional Formatting (3-Color Scale)
    max_row = len(df) + 1
    max_col = len(df.columns) + 1

    if max_row > 1 and max_col > 1:
        start_cell = "B2"
        end_col_letter = ws.cell(row=max_row, column=max_col).column_letter
        end_cell = f"{end_col_letter}{max_row}"
        matrix_range = f"{start_cell}:{end_cell}"

        color_scale = ColorScaleRule(
            start_type="num",
            start_value=low_threshold,
            start_color="FFFFFF",  # White (low)
            mid_type="num",
            mid_value=mid_threshold,
            mid_color="FEF08A",  # Yellow (mid)
            end_type="num",
            end_value=high_threshold,
            end_color="EF4444",  # Red (high)
        )
        ws.conditional_formatting.add(matrix_range, color_scale)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return wb


def export_similarity_matrix_to_excel(
    df: pd.DataFrame, threshold: float = 0.59, write_only: bool = False
) -> bytes:
    """Exports a similarity matrix DataFrame into an in-memory Excel file (.xlsx) with formatting."""
    wb = build_similarity_workbook(
        df,
        threshold=threshold,
        write_only=write_only,
        mid_threshold=threshold,
    )
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_similarity_matrix_to_temp_file(
    df: pd.DataFrame, threshold: float = 0.59, write_only: bool = False
) -> str:
    """
    Exports the similarity matrix to a temporary .xlsx file on disk.
    The created file is automatically registered for cleanup on application exit via atexit.

    Returns:
        str: Absolute path to the created temporary Excel file.
    """
    wb = build_similarity_workbook(
        df,
        threshold=threshold,
        write_only=write_only,
        mid_threshold=threshold,
    )
    temp_path = _create_managed_temp_file(suffix=".xlsx", prefix="similarity_matrix_")
    wb.save(temp_path)
    return temp_path


def generate_csv_matrix_stream(matrix_df: pd.DataFrame) -> Generator[str, None, None]:
    """
    Yields CSV formatted lines line-by-line from a similarity matrix DataFrame.

    Memory-efficient generator for exporting large result sets (>10,000 document pairs)
    without materializing the entire formatted output string or Excel workbook in memory.

    Args:
        matrix_df (pd.DataFrame): Similarity matrix DataFrame with document labels as index and columns.

    Yields:
        str: CSV formatted string row (including newline character).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Yield header row. Column labels are sanitized for the same reason as in
    # build_similarity_workbook(): a CSV opened in Excel evaluates formulas too,
    # so the streaming route must not be a way around the protection.
    header = ["Document"] + [sanitize_spreadsheet_value(c) for c in matrix_df.columns]
    writer.writerow(header)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    # Yield data rows line by line
    for index, row in matrix_df.iterrows():
        writer.writerow(
            [sanitize_spreadsheet_value(index)]
            + [sanitize_spreadsheet_value(v) for v in row.tolist()]
        )
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


def generate_tsv_matrix_stream(matrix_df: pd.DataFrame) -> Generator[str, None, None]:
    """
    Yields TSV formatted lines line-by-line from a similarity matrix DataFrame.

    Memory-efficient generator for exporting large result sets (>10,000 document pairs)
    using tab-delimited formatting for R and Pandas workflows.

    Args:
        matrix_df (pd.DataFrame): Similarity matrix DataFrame with document labels as index and columns.

    Yields:
        str: TSV formatted string row (including newline character).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter="\t")

    # Yield header row
    header = ["Document"] + [sanitize_spreadsheet_value(c) for c in matrix_df.columns]
    writer.writerow(header)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    # Yield data rows line by line
    for index, row in matrix_df.iterrows():
        writer.writerow(
            [sanitize_spreadsheet_value(index)]
            + [sanitize_spreadsheet_value(v) for v in row.tolist()]
        )
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)

